import openpyxl, os
os.chdir('"C:\\Users\\Jeffery John\\PycharmProjects\\Automate the Boring Stuff"')

workbook = openpyxl.load_workbook('example.xslx')
sheet = workbook.get_sheet_by_name('Sheet1')

wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'].value = 42
sheet['A1'].value = 'Hello World'
wb.save('example.xlsx')
openpyxl.load_workbook('')

sheet2 = wb.create_sheet()
sheet2.title = 'Second Sheet'
print(wb.get_sheet_names)